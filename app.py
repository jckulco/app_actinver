import faulthandler
faulthandler.enable()

import io
import json
import pandas as pd
import streamlit as st
import altair as alt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from clean_engine import clean_tenable_export
from openpages_mapper import build_openpages_export

# ---------------------------------------------------------------------------
# Carbon Design System tokens (IBM)
# ---------------------------------------------------------------------------
CARBON_BLUE_60 = "#0f62fe"
CARBON_BLUE_70 = "#0043ce"
CARBON_GRAY_100 = "#161616"
CARBON_GRAY_90 = "#262626"
CARBON_GRAY_20 = "#e0e0e0"
CARBON_GRAY_10 = "#f4f4f4"
CARBON_WHITE = "#ffffff"
CARBON_SUPPORT_ERROR = "#da1e28"
CARBON_SUPPORT_SUCCESS = "#24a148"
CARBON_SUPPORT_WARNING = "#f1c21b"

st.set_page_config(page_title="Limpieza de Vulnerabilidades · Tenable", layout="wide")

CARBON_CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {{
    font-family: 'IBM Plex Sans', sans-serif;
    color: {CARBON_GRAY_100};
}}

.stApp {{
    background-color: {CARBON_WHITE};
}}

/* Carbon-style header */
.carbon-header {{
    background-color: {CARBON_GRAY_100};
    color: {CARBON_WHITE};
    padding: 1.25rem 2rem;
    margin: -6rem -4rem 2rem -4rem;
    border-bottom: 4px solid {CARBON_BLUE_60};
}}
.carbon-header h1 {{
    font-size: 1.25rem;
    font-weight: 500;
    margin: 0;
    letter-spacing: 0;
}}
.carbon-header p {{
    font-size: 0.8rem;
    color: {CARBON_GRAY_20};
    margin: 0.15rem 0 0 0;
}}

/* Kill rounded corners everywhere, Carbon uses 0 border-radius */
div, button, input, textarea, section {{
    border-radius: 0px !important;
}}

/* Buttons */
.stButton>button, .stDownloadButton>button {{
    background-color: {CARBON_BLUE_60};
    color: {CARBON_WHITE};
    border: none;
    font-weight: 500;
    font-size: 0.875rem;
    padding: 0.65rem 1.25rem;
}}
.stButton>button:hover, .stDownloadButton>button:hover {{
    background-color: {CARBON_BLUE_70};
    color: {CARBON_WHITE};
}}

/* Metric cards, Carbon tile style */
[data-testid="stMetric"] {{
    background-color: {CARBON_GRAY_10};
    border-left: 3px solid {CARBON_BLUE_60};
    padding: 1rem 1rem 0.75rem 1rem;
}}
[data-testid="stMetricLabel"] {{
    font-size: 0.75rem;
    color: {CARBON_GRAY_90};
    text-transform: uppercase;
    letter-spacing: 0.02em;
}}

/* Section labels, Carbon "eyebrow" style */
.carbon-eyebrow {{
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: {CARBON_BLUE_60};
    margin-bottom: 0.25rem;
}}

.carbon-rule {{
    border: none;
    border-top: 1px solid {CARBON_GRAY_20};
    margin: 2rem 0 1.5rem 0;
}}

table {{
    font-size: 0.85rem;
}}

.stTabs [data-baseweb="tab-list"] {{
    gap: 0;
    border-bottom: 1px solid {CARBON_GRAY_20};
}}
.stTabs [data-baseweb="tab"] {{
    font-family: 'IBM Plex Sans', sans-serif;
    font-size: 0.875rem;
}}
.stTabs [aria-selected="true"] {{
    border-bottom: 2px solid {CARBON_BLUE_60} !important;
    color: {CARBON_BLUE_60} !important;
}}
</style>
"""
st.markdown(CARBON_CSS, unsafe_allow_html=True)

st.markdown(
    """
    <div class="carbon-header">
        <h1>Limpieza de Vulnerabilidades — Tenable</h1>
        <p>Consolidación de activos duplicados por inconsistencia de sistema operativo</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Reglas aplicadas (visible para usuarios sin conocimientos técnicos)
# ---------------------------------------------------------------------------
with st.expander("Reglas de limpieza aplicadas en esta versión", expanded=False):
    st.markdown(
        """
- **Identidad del activo**: se agrupan las filas por nombre de equipo (`asset.host_name`).
  Si no hay nombre de equipo registrado, se usa la dirección IP (`asset.ipv4_addresses`) en su lugar.
- **Sistema operativo del activo**: cuando el mismo activo tiene más de un valor de sistema
  operativo registrado (por errores de escaneo), se conserva el **más específico**
  (por ejemplo, "Windows Server 2012 R2 Build 9600" gana sobre "Windows Server 2012").
- **Vulnerabilidades duplicadas**: una vulnerabilidad se considera la misma cuando coincide
  el activo, el nombre de la vulnerabilidad y el puerto. Si aparece más de una vez, se
  conserva una sola fila, priorizando la de mayor severidad y la más reciente.
- **Nada se descarta silenciosamente**: cada fila que se fusiona queda registrada en la
  hoja de auditoría del archivo de salida, con el detalle de qué se conservó y por qué.
- **Joya de la Corona**: se compara `asset.ipv4_addresses` contra un listado externo de IPs
  críticas (`listado_joyas.xlsx`, columna `Joyas`). Si hace match, el activo se marca como
  **Joya de la Corona**. Adicionalmente se valida que `asset.tags` contenga la etiqueta
  `01.ACT.JOYAS`: si una IP está en el listado pero el activo **no** trae esa etiqueta, se
  marca como una inconsistencia a revisar (IP crítica sin la clasificación formal esperada).
        """
    )

st.markdown('<hr class="carbon-rule">', unsafe_allow_html=True)
st.markdown('<div class="carbon-eyebrow">Paso 1</div>', unsafe_allow_html=True)
col_a, col_b = st.columns(2)
with col_a:
    uploaded = st.file_uploader("Carga el archivo exportado de Tenable (.xlsx)", type=["xlsx"])
with col_b:
    joyas_uploaded = st.file_uploader(
        "Carga el listado de IPs 'Joya de la Corona' (.xlsx, opcional)", type=["xlsx"]
    )

if uploaded is not None:
    try:
        joyas_df = pd.read_excel(joyas_uploaded) if joyas_uploaded is not None else None
        result = clean_tenable_export(pd.read_excel(uploaded), joyas_df)
    except Exception as e:
        st.error(f"No se pudo procesar el archivo. Detalle: {e}")
        st.stop()

    ins = result.insights

    st.markdown('<hr class="carbon-rule">', unsafe_allow_html=True)
    st.markdown('<div class="carbon-eyebrow">Paso 2 · Resultado</div>', unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Filas originales", ins["filas_originales"])
    c2.metric("Filas limpias", ins["filas_limpias"])
    c3.metric("Duplicados eliminados", ins["filas_eliminadas_por_duplicado"])
    c4.metric("Reducción", f"{ins['pct_reduccion']}%")

    c5, c6, c7 = st.columns(3)
    c5.metric("Activos únicos", ins["activos_unicos"])
    c6.metric("Activos con SO inconsistente", ins["activos_con_multiples_os_registrados"])
    c7.metric("Vulnerabilidades críticas conservadas", ins["severidad_criticas_conservadas"])

    if joyas_uploaded is not None:
        c8, c9 = st.columns(2)
        c8.metric("Activos 'Joya de la Corona'", ins["activos_joya_corona"])
        c9.metric(
            "Joyas SIN tag 01.ACT.JOYAS",
            ins["activos_joya_sin_tag"],
            help="IP está en el listado de Joyas pero asset.tags no trae la etiqueta 01.ACT.JOYAS",
        )
        if ins["activos_joya_sin_tag"] > 0:
            st.warning(
                f"⚠️ {ins['activos_joya_sin_tag']} activo(s) tienen una IP clasificada como "
                "'Joya de la Corona' pero no traen la etiqueta 01.ACT.JOYAS en asset.tags. "
                "Revisar en la pestaña 'Joyas de la Corona'."
            )

    st.markdown('<hr class="carbon-rule">', unsafe_allow_html=True)
    st.markdown('<div class="carbon-eyebrow">Paso 3 · Detalle</div>', unsafe_allow_html=True)

    tab_names = ["Datos limpios", "Auditoría de fusiones", "Distribución de severidad"]
    if joyas_uploaded is not None:
        tab_names.append("Joyas de la Corona")
    tabs = st.tabs(tab_names)
    tab1, tab2, tab3 = tabs[0], tabs[1], tabs[2]

    with tab1:
        st.dataframe(result.clean_df, use_container_width=True, height=420)

    with tab2:
        if result.audit_df.empty:
            st.info("No se encontraron filas duplicadas que requirieran fusión.")
        else:
            st.dataframe(result.audit_df, use_container_width=True, height=420)

    if joyas_uploaded is not None:
        with tabs[3]:
            joyas_rows = result.clean_df[result.clean_df["es_joya_corona"]]
            if joyas_rows.empty:
                st.info("Ninguna IP del archivo coincidió con el listado de Joyas de la Corona.")
            else:
                sin_tag = joyas_rows[~joyas_rows["tiene_tag_joyas"]]
                if not sin_tag.empty:
                    st.markdown("**Inconsistencias — IP en listado de Joyas, sin tag `01.ACT.JOYAS`:**")
                    st.dataframe(
                        sin_tag[
                            ["asset_id_canonical", "asset.host_name", "asset.ipv4_addresses",
                             "asset.tags", "clasificacion_joyas"]
                        ],
                        use_container_width=True,
                    )
                st.markdown("**Todos los activos clasificados como Joya de la Corona:**")
                st.dataframe(
                    joyas_rows[
                        ["asset_id_canonical", "asset.host_name", "asset.ipv4_addresses",
                         "tiene_tag_joyas", "clasificacion_joyas"]
                    ],
                    use_container_width=True,
                )

    with tab3:
        sev_order = ["Critical", "High", "Medium", "Low", "Info"]
        sev_counts = (
            result.clean_df["severity"]
            .value_counts()
            .reindex(sev_order)
            .dropna()
            .reset_index()
        )
        sev_counts.columns = ["severity", "count"]
        chart = (
            alt.Chart(sev_counts)
            .mark_bar()
            .encode(
                x=alt.X("severity", sort=sev_order, title=None),
                y=alt.Y("count", title="Vulnerabilidades"),
                color=alt.Color(
                    "severity",
                    scale=alt.Scale(
                        domain=sev_order,
                        range=[CARBON_SUPPORT_ERROR, "#ff832b", CARBON_SUPPORT_WARNING, CARBON_BLUE_60, CARBON_GRAY_20],
                    ),
                    legend=None,
                ),
            )
            .properties(height=320)
        )
        st.altair_chart(chart, use_container_width=True)

    # -----------------------------------------------------------------
    # Descarga
    # -----------------------------------------------------------------
    st.markdown('<hr class="carbon-rule">', unsafe_allow_html=True)
    st.markdown('<div class="carbon-eyebrow">Paso 4 · Descarga</div>', unsafe_allow_html=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result.clean_df.to_excel(writer, sheet_name="Limpio", index=False)
        result.audit_df.to_excel(writer, sheet_name="Auditoria", index=False)
        pd.DataFrame(list(ins.items()), columns=["Metrica", "Valor"]).to_excel(
            writer, sheet_name="Insights", index=False
        )
    buffer.seek(0)

    wb = load_workbook(buffer)
    header_fill = PatternFill("solid", start_color="0F62FE")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    body_font = Font(name="Arial")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)
        ws.freeze_panes = "A2"

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    st.download_button(
        label="Descargar archivo limpio (.xlsx)",
        data=final_buffer,
        file_name="vulnerabilidades_limpio.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown('<hr class="carbon-rule">', unsafe_allow_html=True)
    st.markdown('<div class="carbon-eyebrow">Paso 5 · Exportar a OpenPages</div>', unsafe_allow_html=True)
    st.caption(
        "Payloads JSON listos para la API v2 de OpenPages (patrón de ejemplo_carga.py). "
        "`type_definition_id` y `primary_parent_id` quedan en `null`: se resuelven en tiempo de "
        "ejecución (`get_all_types`) y con la tabla de correspondencia asset_id_canonical ↔ ID de "
        "OpenPages, aún pendiente. Nombres de campo configurables en `openpages_mapper.py`."
    )
    openpages_export = build_openpages_export(result.clean_df)
    openpages_json = json.dumps(openpages_export, indent=2, default=str, ensure_ascii=False)
    st.download_button(
        label="Descargar payloads OpenPages (.json)",
        data=openpages_json,
        file_name="openpages_payloads.json",
        mime="application/json",
    )
else:
    st.info("Carga un archivo .xlsx exportado de Tenable para comenzar.")
